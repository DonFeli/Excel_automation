from argparse import ArgumentParser
from config import logger
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
import pandas as pd

from processor.extractor.extractor import CollectOperationExtractor
from processor.transformer.collect_transformer import CollectTransformer
from processor.indicator.stopped_collect import StoppedCollectDetector
from config.definitions import SheetParameters

pd.set_option("display.max_columns", 500)
pd.set_option("display.width", 0)


def parse_arguments():
    """Parse argument from console"""
    parser = ArgumentParser()
    parser.add_argument('--gsheet', '-g', type=str, default='Excel automation project', help='Name of excel file.')
    return parser.parse_args()


def print_rows(sheet):
    for row in sheet.iter_rows(values_only=True):
        print(row)


def main(args):
    logger.info('-------------- Retrieving collects --------------------')
    collect_operations = CollectOperationExtractor()
    collect_operations.retrieve_collects()

    # LOAD AND TRANSFORM
    # - Transforme la liste de dictionnaires en dataframe
    # - Extrait du nested dictionnary les champs item scraped count et finish reason
    # - Exporte tel quel dans Excel
    logger.info('-------------- Sheet 1 : Load and transform --------------------')

    structured_collects = CollectTransformer().load_and_transform(collect_operations.collect_data)

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "load_and_transform"
    for row in dataframe_to_rows(structured_collects, index=False, header=True):
        sheet1.append(row)

    # FILTER INSUFFICIENT COLLECTS
    # - Filtre les territoires ayant au moins 2 collectes
    # - Exporte tel quel dans Excel
    logger.info('-------------- Sheet 2 : Filter insufficient collects --------------------')

    sufficient_collects = StoppedCollectDetector.filter_insufficient_collects(structured_collects, 2)

    sufficient_collects_sheet = workbook.create_sheet("filter_insufficient_collects")
    for row in dataframe_to_rows(sufficient_collects, index=False, header=True):
        sufficient_collects_sheet.append(row)

    # PAIRS WITH LAST COLLECTS
    # - Cree un dataframe contenant uniquement les dernieres collectes de chaque territoire
    # - Merge avec l'ancien dataframe pour obtenir pour chaque collecte, la derniere collecte du territoire
    # - 2 nouvelles colonnes : item scraped count last et updated at last
    # - Exporte tel quel dans Excel
    logger.info('-------------- Sheet 3 : Get pairs with last collect --------------------')

    pairs_with_last_collect = StoppedCollectDetector.get_pairs_with_last_collect(sufficient_collects)

    pairs_with_last_collect_sheet = workbook.create_sheet("get_pairs_with_last_collect")
    for row in dataframe_to_rows(pairs_with_last_collect, index=False, header=True):
        pairs_with_last_collect_sheet.append(row)

    # PROCESSED PAIRS WITH EXCEL
    # - Cree 3 nouvelles colonnes dans Excel
    #   - interval : nombre de jours entre une collecte et la derniere collecte d'un territoire
    #   - is stopped : si item scraped count est égal à 0
    #   - was active : si item scraped count est supérieur à 9
    logger.info('-------------- Sheet 4 : Get processed pairs Excel --------------------')

    processed_pairs_excel_sheet = workbook.copy_worksheet(pairs_with_last_collect_sheet)
    processed_pairs_excel_sheet.title = 'get_processed_pairs Excel'

    processed_pairs_excel_sheet['J1'].value = "interval"
    processed_pairs_excel_sheet['J2'] = "=(H2-A2)"
    processed_pairs_excel_sheet['K1'].value = "is_stopped"
    processed_pairs_excel_sheet['K2'] = '=IF(I2=0, "TRUE", "FALSE")'
    processed_pairs_excel_sheet['L1'].value = "was_active"
    processed_pairs_excel_sheet['L2'] = '=IF(E2>9, "TRUE", "FALSE")'

    for i in range(2, processed_pairs_excel_sheet.max_row):
        processed_pairs_excel_sheet[f'J{i}'] = Translator("=(H2-A2)", origin="J2").translate_formula(f'J{i}')
        processed_pairs_excel_sheet[f'K{i}'] = Translator('=IF(I2=0, "TRUE", "FALSE")', origin="K2").translate_formula(
            f'K{i}')
        processed_pairs_excel_sheet[f'L{i}'] = Translator('=IF(E2>9, "TRUE", "FALSE")', origin="L2").translate_formula(
            f'L{i}')

    # PROCESSED PAIRS WITH PYTHON
    # - La meme chose mais transformé dans Python
    logger.info('-------------- Sheet 5 : Get processed pairs Python --------------------')

    processed_pairs = StoppedCollectDetector.get_processed_pairs(pairs_with_last_collect)
    processed_pairs.sort_values(["territory_uid", "updated_at"], inplace=True, ascending=False)

    processed_pairs_python_sheet = workbook.create_sheet("get_processed_pairs Python")
    for row in dataframe_to_rows(processed_pairs, index=False, header=True):
        processed_pairs_python_sheet.append(row)

    # SUM SCRAPED ON CURRENTLY STOPPED COLLECTS
    # - Cree une colonne calculant la somme cumulée des item scraped count par territoire
    # - Exporte tel quel dans Excel
    logger.info('-------------- Sheet 6 : Get sum scraped currently stopped --------------------')

    currently_stopped = StoppedCollectDetector.get_sum_scraped_currently_stopped(processed_pairs)

    currently_stopped_sheet = workbook.create_sheet('get_sum_scraped_currently_stopped')
    for row in dataframe_to_rows(currently_stopped, index=False, header=True):
        currently_stopped_sheet.append(row)

    # EVER ACTIVE EXCEL
    # - Filtre les territoires pour lesquels il y a eu au moins une collecte active (was_active = TRUE)
    # logger.info('-------------- Sheet 7 : Get ever active Excel --------------------')
    #
    # ever_active_sheet_E = workbook.copy_worksheet(currently_stopped_sheet)
    # ever_active_sheet_E.title = 'get_ever_active Excel'
    #
    # # Récupère les territory_uids où la collecte était active (was_active = True)
    # ever_active_rows = []
    # for row in ever_active_sheet_E.iter_rows(min_row=2, max_row=ever_active_sheet_E.max_row, min_col=12, max_col=12):
    #     for cell in row:
    #         if cell.value == True:
    #             ever_active_rows.append(row[0].row)
    #
    # ever_active_uids = []
    # for active_row in ever_active_rows:
    #     ever_active_uids.append(ever_active_sheet_E[f'C{active_row}'].value)
    #
    # # Supprime les collectes de territoires qui n'ont jamais été actifs
    # for cell in ever_active_sheet_E['C']:
    #     if cell.row != 1:
    #         if cell.value not in ever_active_uids:
    #             ever_active_sheet_E.delete_rows(cell.row)

    # EVER ACTIVE PYTHON
    # - Filtre les territoires pour lesquels il y a eu au moins une collecte active (was_active = TRUE)
    logger.info('-------------- Sheet 7 : Get ever active Python --------------------')
    
    ever_active_uids = StoppedCollectDetector.get_ever_active(currently_stopped)

    ever_active_sheet_P = workbook.create_sheet('get_ever_active Python')
    for row in dataframe_to_rows(ever_active_uids, index=False, header=True):
        ever_active_sheet_P.append(row)

    # GET CANDIDATE EXCEL
    logger.info('-------------- Sheet 8 : Get candidate Excel --------------------')

    get_candidate_sheet_E = workbook.copy_worksheet(ever_active_sheet_P)
    get_candidate_sheet_E.title = 'get_candidate Excel'

    # Supprime les collectes de territoires qui n'ont jamais été actifs
    for cell in get_candidate_sheet_E['M']:
        if cell.row != 1:
            if cell.value != 0:
                get_candidate_sheet_E.delete_rows(cell.row)

    # Supprime les lignes où l'interval est inférieur à 12 jours
    for cell in get_candidate_sheet_E['J']:
        if cell.row != 1:
            if cell.value < 12:
                get_candidate_sheet_E.delete_rows(cell.row)
                
    # GET CANDIDATE PYTHON
    logger.info('-------------- Sheet 8 : Get candidate Python --------------------')
    
    candidates = StoppedCollectDetector.get_candidate(ever_active_uids, currently_stopped, last_active_day_treshold=12)
    
    get_candidate_sheet_P = workbook.create_sheet('get_candidate Python')
    for row in dataframe_to_rows(candidates, index=False, header=True):
        get_candidate_sheet_P.append(row)

    # SUMMARY STATISTICS
    logger.info('-------------- Sheet 9 : Summary --------------------')

    summary_statistics_sheet = workbook.create_sheet("Summary Statistics")
    summary_statistics_sheet['A1'] = '=COUNTA(N3:N236)'
    summary_statistics_sheet['A2'] = '=UNIQUE(C2:C1016)'

    # MISE EN FORME
    logger.info('-------------- Saving Excel file ---------------------------------')
    for sheet in workbook._sheets:
        sheet.column_dimensions[
            get_column_letter(SheetParameters.UpdatedAt.column)].width = SheetParameters.UpdatedAt.width

        sheet.column_dimensions[get_column_letter(SheetParameters.Website.column)].width = SheetParameters.Website.width
        sheet.column_dimensions[
            get_column_letter(SheetParameters.TerritoryUid.column)].width = SheetParameters.TerritoryUid.width
        sheet.column_dimensions[get_column_letter(SheetParameters.Id.column)].width = SheetParameters.Id.width
        sheet.column_dimensions[
            get_column_letter(SheetParameters.ItemScrapedCount.column)].width = SheetParameters.ItemScrapedCount.width
        sheet.column_dimensions[
            get_column_letter(SheetParameters.FinishReason.column)].width = SheetParameters.FinishReason.width
        sheet.column_dimensions[get_column_letter(SheetParameters.Status.column)].width = SheetParameters.Status.width
        sheet.column_dimensions[
            get_column_letter(SheetParameters.UpdatedAtLast.column)].width = SheetParameters.UpdatedAtLast.width
        sheet.column_dimensions[get_column_letter(
            SheetParameters.ItemScrapedCountLast.column)].width = SheetParameters.ItemScrapedCountLast.width
        sheet.column_dimensions[get_column_letter(
            SheetParameters.Interval.column)].width = SheetParameters.Interval.width
        sheet.column_dimensions[get_column_letter(
            SheetParameters.IsStopped.column)].width = SheetParameters.IsStopped.width
        sheet.column_dimensions[get_column_letter(
            SheetParameters.WasActive.column)].width = SheetParameters.WasActive.width
        sheet.column_dimensions[get_column_letter(
            SheetParameters.SumScraped.column)].width = SheetParameters.SumScraped.width

        sheet.auto_filter.ref = sheet.dimensions

    workbook.save(filename=f'{args.gsheet}.xlsx')


if __name__ == '__main__':
    args = parse_arguments()
    main(args)

